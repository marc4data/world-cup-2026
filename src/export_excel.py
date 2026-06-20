"""Export each database table to an Excel workbook — one sheet per table.

Writes the first ``row_limit`` rows (default 10,000) of every table in
``data/worldcup.db`` to its own sheet, named after the table. Useful for quick
eyeballing of the data in Excel/Sheets without writing SQL.

CLI:
    python src/export_excel.py                       # -> reports/worldcup_tables.xlsx
    python src/export_excel.py --out /tmp/wc.xlsx --limit 500
"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import DB_PATH, REPO_ROOT

ROW_LIMIT = 10_000
DEFAULT_OUT = REPO_ROOT / "reports" / "worldcup_tables.xlsx"


def list_tables(conn: sqlite3.Connection) -> list[str]:
    """Every user table AND view in the DB (excludes SQLite internals).

    Schema-driven on purpose: any new table or view added later is exported
    automatically, no change needed here.
    """
    return [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type IN ('table', 'view') "
        "AND name NOT LIKE 'sqlite_%' ORDER BY name"
    )]


def export_tables_to_excel(
    db_path=DB_PATH, out_path=DEFAULT_OUT, *, row_limit: int = ROW_LIMIT
) -> tuple[Path, list[tuple[str, int]]]:
    """Write each table (first ``row_limit`` rows) to its own sheet.

    Returns (output path, [(table, rows_written), ...]).
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        tables = list_tables(conn)
        summary: list[tuple[str, int]] = []
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            for table in tables:
                df = pd.read_sql(f"SELECT * FROM {table} LIMIT {int(row_limit)}", conn)
                sheet = table[:31]  # Excel sheet-name limit
                df.to_excel(writer, sheet_name=sheet, index=False)
                _format_sheet(writer.sheets[sheet], df)
                summary.append((table, len(df)))
        return out_path, summary
    finally:
        conn.close()


_LINK_FONT = Font(color="0563C1", underline="single")


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Freeze + bold the header, add an autofilter, size columns, and make URL
    columns (header ends in 'url') real clickable hyperlinks."""
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(df.columns, start=1):
        sample = (str(v) for v in df[col].head(200))
        width = max([len(str(col))] + [len(s) for s in sample], default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 40)
        if str(col).lower().endswith(("url", "api")):   # *_url + the ESPN/FIFA *_api endpoints
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=i)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = _LINK_FONT


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Export DB tables to Excel (one sheet per table)")
    ap.add_argument("--db", default=str(DB_PATH))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--limit", type=int, default=ROW_LIMIT, help="max rows per table")
    args = ap.parse_args(argv)

    out, summary = export_tables_to_excel(args.db, args.out, row_limit=args.limit)
    print(f"wrote {out}  ({len(summary)} sheets)")
    for table, n in summary:
        print(f"  {table:12s} {n:>6d} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
